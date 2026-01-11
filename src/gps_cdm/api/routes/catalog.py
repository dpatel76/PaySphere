"""
CDM Catalog API Routes.

Provides endpoints for browsing, searching, and editing the CDM data catalog.
"""

import json
import io
import os
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Query, Response
from fastapi.responses import StreamingResponse
import psycopg2
from psycopg2.extras import RealDictCursor

from gps_cdm.api.models.catalog import (
    AllowedValue,
    LegacyMapping,
    CatalogElement,
    CatalogElementUpdate,
    CatalogTableSummary,
    CatalogStats,
    CatalogSearchResponse,
)

router = APIRouter(prefix="/catalog", tags=["catalog"])

# Database configuration
DB_CONFIG = {
    'host': os.environ.get('POSTGRES_HOST', 'localhost'),
    'port': int(os.environ.get('POSTGRES_PORT', 5433)),
    'database': os.environ.get('POSTGRES_DB', 'gps_cdm'),
    'user': os.environ.get('POSTGRES_USER', 'gps_cdm_svc'),
    'password': os.environ.get('POSTGRES_PASSWORD', 'gps_cdm_password'),
}

# Table display names and ordering
CDM_TABLE_CONFIG = {
    # Core payment tables
    'cdm_payment_instruction': {'display_name': 'Payment Instruction', 'order': 1},
    'cdm_payment_identifiers': {'display_name': 'Payment Identifiers', 'order': 2},

    # Party tables
    'cdm_party': {'display_name': 'Party', 'order': 10},
    'cdm_party_identifiers': {'display_name': 'Party Identifiers', 'order': 11},

    # Financial Institution tables
    'cdm_financial_institution': {'display_name': 'Financial Institution', 'order': 20},
    'cdm_institution_identifiers': {'display_name': 'Institution Identifiers', 'order': 21},

    # Account tables
    'cdm_account': {'display_name': 'Account', 'order': 30},
    'cdm_account_identifiers': {'display_name': 'Account Identifiers', 'order': 31},

    # Related tables
    'cdm_charge': {'display_name': 'Charge', 'order': 40},
    'cdm_remittance_information': {'display_name': 'Remittance Information', 'order': 41},
    'cdm_regulatory_reporting': {'display_name': 'Regulatory Reporting', 'order': 42},
    'cdm_settlement': {'display_name': 'Settlement', 'order': 43},
    'cdm_payment_status': {'display_name': 'Payment Status', 'order': 44},
    'cdm_payment_event': {'display_name': 'Payment Event', 'order': 45},

    # Statement tables
    'cdm_account_statement': {'display_name': 'Account Statement', 'order': 50},
    'cdm_statement_extension': {'display_name': 'Statement Extension', 'order': 51},

    # ISO 20022 semantic tables
    'cdm_pacs_fi_customer_credit_transfer': {'display_name': 'pacs.008 FI-Customer Credit Transfer', 'order': 60},
    'cdm_pacs_fi_credit_transfer': {'display_name': 'pacs.009 FI Credit Transfer', 'order': 61},
    'cdm_pacs_fi_direct_debit': {'display_name': 'pacs.003 FI Direct Debit', 'order': 62},
    'cdm_pacs_payment_return': {'display_name': 'pacs.004 Payment Return', 'order': 63},
    'cdm_pacs_fi_payment_status_report': {'display_name': 'pacs.002 Payment Status Report', 'order': 64},
    'cdm_pain_customer_credit_transfer_initiation': {'display_name': 'pain.001 Credit Transfer Initiation', 'order': 70},
    'cdm_pain_customer_direct_debit_initiation': {'display_name': 'pain.008 Direct Debit Initiation', 'order': 71},
    'cdm_pain_customer_payment_status_report': {'display_name': 'pain.002 Payment Status Report', 'order': 72},
    'cdm_camt_bank_to_customer_statement': {'display_name': 'camt.053 Bank-to-Customer Statement', 'order': 80},
    'cdm_camt_bank_to_customer_account_report': {'display_name': 'camt.052 Account Report', 'order': 81},
    'cdm_camt_bank_to_customer_debit_credit_notification': {'display_name': 'camt.054 Debit/Credit Notification', 'order': 82},
    'cdm_camt_fi_payment_cancellation_request': {'display_name': 'camt.056 Payment Cancellation Request', 'order': 83},

    # Extension tables (regional payment schemes)
    'cdm_payment_extension_ach': {'display_name': 'Extension: ACH', 'order': 100},
    'cdm_payment_extension_fedwire': {'display_name': 'Extension: Fedwire', 'order': 101},
    'cdm_payment_extension_chips': {'display_name': 'Extension: CHIPS', 'order': 102},
    'cdm_payment_extension_fednow': {'display_name': 'Extension: FedNow', 'order': 103},
    'cdm_payment_extension_rtp': {'display_name': 'Extension: RTP', 'order': 104},
    'cdm_payment_extension_sepa': {'display_name': 'Extension: SEPA', 'order': 110},
    'cdm_payment_extension_target2': {'display_name': 'Extension: TARGET2', 'order': 111},
    'cdm_payment_extension_chaps': {'display_name': 'Extension: CHAPS', 'order': 112},
    'cdm_payment_extension_fps': {'display_name': 'Extension: FPS', 'order': 113},
    'cdm_payment_extension_bacs': {'display_name': 'Extension: BACS', 'order': 114},
    'cdm_payment_extension_npp': {'display_name': 'Extension: NPP', 'order': 120},
    'cdm_payment_extension_pix': {'display_name': 'Extension: PIX', 'order': 121},
    'cdm_payment_extension_upi': {'display_name': 'Extension: UPI', 'order': 122},
    'cdm_payment_extension_cnaps': {'display_name': 'Extension: CNAPS', 'order': 123},
    'cdm_payment_extension_bojnet': {'display_name': 'Extension: BOJ-NET', 'order': 124},
    'cdm_payment_extension_kftc': {'display_name': 'Extension: KFTC', 'order': 125},
    'cdm_payment_extension_meps_plus': {'display_name': 'Extension: MEPS+', 'order': 126},
    'cdm_payment_extension_rtgs_hk': {'display_name': 'Extension: RTGS HK', 'order': 127},
    'cdm_payment_extension_sarie': {'display_name': 'Extension: SARIE', 'order': 128},
    'cdm_payment_extension_uaefts': {'display_name': 'Extension: UAEFTS', 'order': 129},
    'cdm_payment_extension_promptpay': {'display_name': 'Extension: PromptPay', 'order': 130},
    'cdm_payment_extension_paynow': {'display_name': 'Extension: PayNow', 'order': 131},
    'cdm_payment_extension_instapay': {'display_name': 'Extension: InstaPay', 'order': 132},
    'cdm_payment_extension_swift': {'display_name': 'Extension: SWIFT MT', 'order': 140},
    'cdm_payment_extension_iso20022': {'display_name': 'Extension: ISO 20022 (Generic)', 'order': 150},

    # Other tables
    'cdm_document_identifier': {'display_name': 'Document Identifier', 'order': 200},
    'cdm_identifier_type': {'display_name': 'Identifier Type', 'order': 201},
    'cdm_transaction_identifier': {'display_name': 'Transaction Identifier', 'order': 202},
}


def get_db_connection():
    """Create database connection."""
    return psycopg2.connect(**DB_CONFIG)


def parse_allowed_values(data) -> Optional[List[AllowedValue]]:
    """Parse allowed values from JSON or None."""
    if data is None:
        return None
    if isinstance(data, str):
        data = json.loads(data)
    if isinstance(data, list):
        return [AllowedValue(**v) if isinstance(v, dict) else AllowedValue(value=str(v)) for v in data]
    return None


def parse_legacy_mappings(data) -> List[LegacyMapping]:
    """Parse legacy mappings from JSON."""
    if data is None:
        return []
    if isinstance(data, str):
        data = json.loads(data)
    if isinstance(data, list):
        # Filter out empty mappings
        return [LegacyMapping(**m) for m in data if m.get('format_id')]
    return []


@router.get("/stats", response_model=CatalogStats)
async def get_catalog_stats():
    """Get overall catalog statistics."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # Total elements
            cur.execute("""
                SELECT
                    COUNT(*) as total_elements,
                    COUNT(*) FILTER (WHERE business_description IS NOT NULL AND business_description <> '') as documented_elements,
                    COUNT(*) FILTER (WHERE iso_element_name IS NOT NULL AND iso_element_name <> '') as iso_mapped_elements,
                    COUNT(DISTINCT pde_table_name) as total_tables
                FROM mapping.cdm_catalog
                WHERE is_active = true
            """)
            row = cur.fetchone()
            total = row[0] or 0
            documented = row[1] or 0
            iso_mapped = row[2] or 0
            total_tables = row[3] or 0

        conn.close()

        return CatalogStats(
            total_elements=total,
            documented_elements=documented,
            iso_mapped_elements=iso_mapped,
            total_tables=total_tables,
            documentation_pct=round(100 * documented / max(total, 1), 1),
            iso_mapping_pct=round(100 * iso_mapped / max(total, 1), 1),
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tables", response_model=List[CatalogTableSummary])
async def get_catalog_tables():
    """Get list of CDM tables with element counts."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("""
                SELECT
                    pde_table_name,
                    COUNT(*) as element_count,
                    COUNT(*) FILTER (WHERE business_description IS NOT NULL AND business_description <> '') as documented_count,
                    COUNT(*) FILTER (WHERE iso_element_name IS NOT NULL AND iso_element_name <> '') as iso_mapped_count,
                    ROUND(100.0 * COUNT(*) FILTER (WHERE business_description IS NOT NULL AND business_description <> '') / COUNT(*), 1) as documentation_pct
                FROM mapping.cdm_catalog
                WHERE is_active = true
                GROUP BY pde_table_name
                ORDER BY pde_table_name
            """)
            rows = cur.fetchall()
        conn.close()

        tables = []
        for row in rows:
            table_name = row[0]
            config = CDM_TABLE_CONFIG.get(table_name, {
                'display_name': table_name.replace('cdm_', '').replace('_', ' ').title(),
                'order': 999
            })
            tables.append(CatalogTableSummary(
                table_name=table_name,
                display_name=config['display_name'],
                element_count=row[1],
                documented_count=row[2],
                iso_mapped_count=row[3],
                documentation_pct=float(row[4]) if row[4] else 0.0,
                sort_order=config['order'],
            ))

        # Sort by configured order
        tables.sort(key=lambda t: t.sort_order)
        return tables

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/elements", response_model=CatalogSearchResponse)
async def search_catalog_elements(
    query: Optional[str] = Query(None, description="Search query (wildcards supported)"),
    table_name: Optional[str] = Query(None, description="Filter by table name"),
    documented_only: bool = Query(False, description="Only return documented elements"),
    limit: int = Query(500, ge=1, le=5000),
    offset: int = Query(0, ge=0),
):
    """Search catalog elements with optional filtering."""
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            # Build query
            conditions = ["c.is_active = true"]
            params = []

            if table_name:
                conditions.append("c.pde_table_name = %s")
                params.append(table_name)

            if documented_only:
                conditions.append("c.business_description IS NOT NULL AND c.business_description <> ''")

            if query:
                # Support wildcard search
                search_pattern = query.replace('*', '%').replace('?', '_')
                if '%' not in search_pattern and '_' not in search_pattern:
                    search_pattern = f'%{search_pattern}%'

                conditions.append("""(
                    c.business_name ILIKE %s OR
                    c.business_description ILIKE %s OR
                    c.pde_column_name ILIKE %s OR
                    c.iso_element_name ILIKE %s OR
                    c.iso_element_path ILIKE %s
                )""")
                params.extend([search_pattern] * 5)

            where_clause = " AND ".join(conditions)

            # Count total
            count_sql = f"""
                SELECT COUNT(*)
                FROM mapping.cdm_catalog c
                WHERE {where_clause}
            """
            cur.execute(count_sql, params)
            total_count = cur.fetchone()['count']

            # Fetch elements using the view
            fetch_sql = f"""
                SELECT
                    c.catalog_id,
                    c.pde_table_name,
                    c.pde_column_name,
                    c.pde_data_type,
                    c.pde_is_nullable,
                    c.business_name,
                    c.business_description,
                    c.data_format,
                    c.allowed_values,
                    c.iso_element_name,
                    c.iso_element_description,
                    c.iso_element_path,
                    c.iso_data_type,
                    c.updated_at,
                    c.updated_by,
                    COALESCE(
                        (SELECT jsonb_agg(
                            jsonb_build_object(
                                'format_id', gfm.format_id,
                                'source_expression', gfm.source_expression,
                                'transform_expression', gfm.transform_expression,
                                'entity_role', gfm.entity_role
                            )
                        )
                        FROM mapping.gold_field_mappings gfm
                        WHERE gfm.gold_table = c.pde_table_name
                          AND gfm.gold_column = c.pde_column_name
                          AND gfm.is_active = true
                        ), '[]'::jsonb
                    ) as legacy_mappings,
                    (SELECT COUNT(DISTINCT gfm.format_id)
                     FROM mapping.gold_field_mappings gfm
                     WHERE gfm.gold_table = c.pde_table_name
                       AND gfm.gold_column = c.pde_column_name
                       AND gfm.is_active = true
                    ) as legacy_mapping_count
                FROM mapping.cdm_catalog c
                WHERE {where_clause}
                ORDER BY c.pde_table_name, c.pde_column_name
                LIMIT %s OFFSET %s
            """
            params.extend([limit, offset])
            cur.execute(fetch_sql, params)
            rows = cur.fetchall()

        conn.close()

        elements = []
        for row in rows:
            elements.append(CatalogElement(
                catalog_id=row['catalog_id'],
                pde_table_name=row['pde_table_name'],
                pde_column_name=row['pde_column_name'],
                pde_data_type=row['pde_data_type'],
                pde_is_nullable=row['pde_is_nullable'],
                business_name=row['business_name'],
                business_description=row['business_description'],
                data_format=row['data_format'],
                allowed_values=parse_allowed_values(row['allowed_values']),
                iso_element_name=row['iso_element_name'],
                iso_element_description=row['iso_element_description'],
                iso_element_path=row['iso_element_path'],
                iso_data_type=row['iso_data_type'],
                legacy_mappings=parse_legacy_mappings(row['legacy_mappings']),
                legacy_mapping_count=row['legacy_mapping_count'],
                updated_at=row['updated_at'],
                updated_by=row['updated_by'],
            ))

        return CatalogSearchResponse(
            elements=elements,
            total_count=total_count,
            limit=limit,
            offset=offset,
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/elements/{catalog_id}", response_model=CatalogElement)
async def get_catalog_element(catalog_id: int):
    """Get a single catalog element by ID."""
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            cur.execute("""
                SELECT
                    c.catalog_id,
                    c.pde_table_name,
                    c.pde_column_name,
                    c.pde_data_type,
                    c.pde_is_nullable,
                    c.business_name,
                    c.business_description,
                    c.data_format,
                    c.allowed_values,
                    c.iso_element_name,
                    c.iso_element_description,
                    c.iso_element_path,
                    c.iso_data_type,
                    c.updated_at,
                    c.updated_by,
                    COALESCE(
                        (SELECT jsonb_agg(
                            jsonb_build_object(
                                'format_id', gfm.format_id,
                                'source_expression', gfm.source_expression,
                                'transform_expression', gfm.transform_expression,
                                'entity_role', gfm.entity_role
                            )
                        )
                        FROM mapping.gold_field_mappings gfm
                        WHERE gfm.gold_table = c.pde_table_name
                          AND gfm.gold_column = c.pde_column_name
                          AND gfm.is_active = true
                        ), '[]'::jsonb
                    ) as legacy_mappings,
                    (SELECT COUNT(DISTINCT gfm.format_id)
                     FROM mapping.gold_field_mappings gfm
                     WHERE gfm.gold_table = c.pde_table_name
                       AND gfm.gold_column = c.pde_column_name
                       AND gfm.is_active = true
                    ) as legacy_mapping_count
                FROM mapping.cdm_catalog c
                WHERE c.catalog_id = %s AND c.is_active = true
            """, [catalog_id])
            row = cur.fetchone()

        conn.close()

        if not row:
            raise HTTPException(status_code=404, detail=f"Catalog element {catalog_id} not found")

        return CatalogElement(
            catalog_id=row['catalog_id'],
            pde_table_name=row['pde_table_name'],
            pde_column_name=row['pde_column_name'],
            pde_data_type=row['pde_data_type'],
            pde_is_nullable=row['pde_is_nullable'],
            business_name=row['business_name'],
            business_description=row['business_description'],
            data_format=row['data_format'],
            allowed_values=parse_allowed_values(row['allowed_values']),
            iso_element_name=row['iso_element_name'],
            iso_element_description=row['iso_element_description'],
            iso_element_path=row['iso_element_path'],
            iso_data_type=row['iso_data_type'],
            legacy_mappings=parse_legacy_mappings(row['legacy_mappings']),
            legacy_mapping_count=row['legacy_mapping_count'],
            updated_at=row['updated_at'],
            updated_by=row['updated_by'],
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/elements/{catalog_id}", response_model=CatalogElement)
async def update_catalog_element(catalog_id: int, update: CatalogElementUpdate):
    """Update a catalog element's business metadata."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            # Build update statement
            update_fields = []
            values = []

            if update.business_name is not None:
                update_fields.append("business_name = %s")
                values.append(update.business_name if update.business_name else None)

            if update.business_description is not None:
                update_fields.append("business_description = %s")
                values.append(update.business_description if update.business_description else None)

            if update.data_format is not None:
                update_fields.append("data_format = %s")
                values.append(update.data_format if update.data_format else None)

            if update.allowed_values is not None:
                update_fields.append("allowed_values = %s")
                values.append(json.dumps([v.model_dump() for v in update.allowed_values]) if update.allowed_values else None)

            if update.iso_element_name is not None:
                update_fields.append("iso_element_name = %s")
                values.append(update.iso_element_name if update.iso_element_name else None)

            if update.iso_element_description is not None:
                update_fields.append("iso_element_description = %s")
                values.append(update.iso_element_description if update.iso_element_description else None)

            if update.iso_element_path is not None:
                update_fields.append("iso_element_path = %s")
                values.append(update.iso_element_path if update.iso_element_path else None)

            if update.iso_data_type is not None:
                update_fields.append("iso_data_type = %s")
                values.append(update.iso_data_type if update.iso_data_type else None)

            if not update_fields:
                raise HTTPException(status_code=400, detail="No fields to update")

            update_fields.append("updated_at = CURRENT_TIMESTAMP")
            update_fields.append("updated_by = %s")
            values.append('api')

            sql = f"""
                UPDATE mapping.cdm_catalog
                SET {', '.join(update_fields)}
                WHERE catalog_id = %s AND is_active = true
                RETURNING catalog_id
            """
            values.append(catalog_id)

            cur.execute(sql, values)
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail=f"Catalog element {catalog_id} not found")

            conn.commit()

        conn.close()

        # Return updated element
        return await get_catalog_element(catalog_id)

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/refresh")
async def refresh_catalog_from_schema():
    """Refresh the catalog from the database schema."""
    try:
        conn = get_db_connection()
        with conn.cursor() as cur:
            cur.execute("SELECT * FROM mapping.refresh_cdm_catalog_from_schema()")
            result = cur.fetchone()
            conn.commit()
        conn.close()

        return {
            "status": "success",
            "inserted": result[0] if result else 0,
            "updated": result[1] if result else 0,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/export")
async def export_catalog(
    table_names: Optional[str] = Query(None, description="Comma-separated table names to export"),
    include_legacy_mappings: bool = Query(True, description="Include legacy format mappings"),
    format: str = Query("csv", description="Export format: csv or xlsx"),
):
    """Export catalog to CSV or Excel format."""
    try:
        conn = get_db_connection()
        with conn.cursor(cursor_factory=RealDictCursor) as cur:
            conditions = ["c.is_active = true"]
            params = []

            if table_names:
                table_list = [t.strip() for t in table_names.split(',')]
                placeholders = ','.join(['%s'] * len(table_list))
                conditions.append(f"c.pde_table_name IN ({placeholders})")
                params.extend(table_list)

            where_clause = " AND ".join(conditions)

            sql = f"""
                SELECT
                    c.pde_table_name as "Table Name",
                    c.pde_column_name as "Column Name",
                    c.pde_data_type as "Data Type",
                    c.pde_is_nullable as "Nullable",
                    c.business_name as "Business Name",
                    c.business_description as "Business Description",
                    c.data_format as "Format",
                    c.iso_element_name as "ISO Element Name",
                    c.iso_element_path as "ISO Element Path",
                    c.iso_data_type as "ISO Data Type",
                    (SELECT COUNT(DISTINCT gfm.format_id)
                     FROM mapping.gold_field_mappings gfm
                     WHERE gfm.gold_table = c.pde_table_name
                       AND gfm.gold_column = c.pde_column_name
                       AND gfm.is_active = true
                    ) as "Legacy Format Count"
                FROM mapping.cdm_catalog c
                WHERE {where_clause}
                ORDER BY c.pde_table_name, c.pde_column_name
            """
            cur.execute(sql, params)
            rows = cur.fetchall()

        conn.close()

        if format == "csv":
            # Generate CSV
            import csv
            output = io.StringIO()
            if rows:
                writer = csv.DictWriter(output, fieldnames=rows[0].keys())
                writer.writeheader()
                for row in rows:
                    writer.writerow(row)

            return Response(
                content=output.getvalue(),
                media_type="text/csv",
                headers={"Content-Disposition": "attachment; filename=cdm_catalog.csv"}
            )
        else:
            # Generate Excel
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
            except ImportError:
                raise HTTPException(status_code=500, detail="openpyxl not installed for Excel export")

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "CDM Catalog"

            # Header row
            if rows:
                headers = list(rows[0].keys())
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")

                # Data rows
                for row_num, row in enumerate(rows, 2):
                    for col, value in enumerate(row.values(), 1):
                        ws.cell(row=row_num, column=col, value=value)

                # Auto-size columns
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter
                    for cell in col:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column].width = adjusted_width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return StreamingResponse(
                output,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=cdm_catalog.xlsx"}
            )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
